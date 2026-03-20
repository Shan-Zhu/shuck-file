"""Excel (.xlsx) extractor for shuck-file."""

import sys
from pathlib import Path
from .base import BaseExtractor


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).replace("|", "\\|").replace("\n", " ")


class XlsxExtractor(BaseExtractor):

    def extract(self, filepath: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("Missing dependency: openpyxl\n  pip install openpyxl", file=sys.stderr)
            sys.exit(1)

        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            while rows and all(c is None for c in rows[-1]):
                rows.pop()
            if not rows:
                sections.append(f"## Sheet: {sheet_name}\n\n*Empty sheet*\n")
                continue

            max_col = 0
            for row in rows:
                for i in range(len(row) - 1, -1, -1):
                    if row[i] is not None:
                        max_col = max(max_col, i + 1)
                        break
            if max_col == 0:
                sections.append(f"## Sheet: {sheet_name}\n\n*Empty sheet*\n")
                continue

            trimmed = [row[:max_col] for row in rows]

            md_lines = [f"## Sheet: {sheet_name}\n"]
            header = trimmed[0]
            md_lines.append("| " + " | ".join(_cell_str(c) for c in header) + " |")
            md_lines.append("| " + " | ".join(["---"] * max_col) + " |")
            for row in trimmed[1:]:
                md_lines.append("| " + " | ".join(_cell_str(c) for c in row) + " |")
            md_lines.append("")
            sections.append("\n".join(md_lines))

        wb.close()
        return "\n".join(sections).strip() + "\n"

    def estimate_tokens(self, filepath: Path) -> int:
        return filepath.stat().st_size // 3

    def extract_tables(self, filepath: Path) -> str:
        # For xlsx, all content IS tables
        return self.extract(filepath)

    def extract_schema(self, filepath: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ""

        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                sections.append(f"## Sheet: {sheet_name}\n\n*Empty sheet*\n")
                continue

            # Get header row and infer types from first data row
            header = rows[0]
            max_col = 0
            for i in range(len(header) - 1, -1, -1):
                if header[i] is not None:
                    max_col = i + 1
                    break
            if max_col == 0:
                sections.append(f"## Sheet: {sheet_name}\n\n*Empty sheet*\n")
                continue

            header = header[:max_col]
            # Infer types from first non-empty data row
            types = ["text"] * max_col
            for row in rows[1:]:
                row = row[:max_col]
                for i, cell in enumerate(row):
                    if cell is not None:
                        if isinstance(cell, (int, float)):
                            types[i] = "number"
                        elif isinstance(cell, bool):
                            types[i] = "boolean"
                        else:
                            types[i] = "text"
                break

            data_rows = sum(1 for r in rows[1:] if any(c is not None for c in r))
            md_lines = [f"## Sheet: {sheet_name} ({data_rows} rows)\n"]
            md_lines.append("| Column | Type |")
            md_lines.append("| --- | --- |")
            for col, typ in zip(header, types):
                md_lines.append(f"| {_cell_str(col)} | {typ} |")
            md_lines.append("")
            sections.append("\n".join(md_lines))

        wb.close()
        return "\n".join(sections).strip() + "\n"

    def extract_sample(self, filepath: Path, n: int = 5) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ""

        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            while rows and all(c is None for c in rows[-1]):
                rows.pop()
            if not rows:
                sections.append(f"## Sheet: {sheet_name}\n\n*Empty sheet*\n")
                continue

            max_col = 0
            for row in rows:
                for i in range(len(row) - 1, -1, -1):
                    if row[i] is not None:
                        max_col = max(max_col, i + 1)
                        break
            if max_col == 0:
                continue

            # Header + first n rows
            sample_rows = [row[:max_col] for row in rows[:n + 1]]
            total = len(rows) - 1

            md_lines = [f"## Sheet: {sheet_name} (showing {min(n, total)}/{total} rows)\n"]
            md_lines.append("| " + " | ".join(_cell_str(c) for c in sample_rows[0]) + " |")
            md_lines.append("| " + " | ".join(["---"] * max_col) + " |")
            for row in sample_rows[1:]:
                md_lines.append("| " + " | ".join(_cell_str(c) for c in row) + " |")
            md_lines.append("")
            sections.append("\n".join(md_lines))

        wb.close()
        return "\n".join(sections).strip() + "\n"
